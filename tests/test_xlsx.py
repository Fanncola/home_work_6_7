import os.path

from openpyxl import load_workbook
from conftest import RES_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_open_xlsx_file():
    file_name = 'file_example_XLSX_50.xlsx'
    workbook = load_workbook(os.path.join(RES_PATH, file_name))
    sheet = workbook.active

    assert len(sheet.tables) == 0
    assert sheet.cell(row=3, column=2).value == 'Mara'
